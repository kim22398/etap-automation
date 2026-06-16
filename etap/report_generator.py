"""
Excel equipment report generator.

Produces a formatted .xlsx workbook with one sheet per equipment category
(Cables, Breakers, Transformers) plus a summary sheet.

Requires: openpyxl
"""

from __future__ import annotations

import datetime
from pathlib import Path
from typing import Any, Union

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from etap.equipment import Cable, Breaker, Transformer


# ---------------------------------------------------------------------------
# Colour palette (ARGB hex strings for openpyxl)
# ---------------------------------------------------------------------------
DARK_BLUE   = "FF1F3864"
MID_BLUE    = "FF2E75B6"
LIGHT_BLUE  = "FFDAE3F3"
HEADER_BG   = "FF1F3864"
ALT_ROW_BG  = "FFEEF3FB"
WHITE       = "FFFFFFFF"
LIGHT_GREY  = "FFD9D9D9"
WARN_ORANGE = "FFFF8C00"
PASS_GREEN  = "FF70AD47"


def _thin_border() -> Border:
    s = Side(style="thin", color="FFC0C0C0")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_style(ws, row: int, cols: int, title: str) -> None:
    """Merge cells across the row and apply the title style."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1)
    cell.value = title
    cell.font = Font(bold=True, size=14, color=WHITE, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22


def _col_header_row(ws, row: int, headers: list[str]) -> None:
    """Write column headers with dark blue background."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True, color=WHITE, name="Calibri", size=10)
        cell.fill = PatternFill("solid", fgColor=MID_BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()
    ws.row_dimensions[row].height = 32


def _data_row(ws, row: int, values: list[Any], alt: bool = False) -> None:
    """Write a data row with optional alternating row colour."""
    bg = ALT_ROW_BG if alt else WHITE
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = Font(name="Calibri", size=10)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()


def _auto_width(ws, min_width: int = 10, max_width: int = 35) -> None:
    """Heuristic column width sizing."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        length = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        ws.column_dimensions[col_letter].width = min(max(length + 2, min_width), max_width)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_cable_sheet(wb: Workbook, cables: list[Cable]) -> None:
    ws = wb.create_sheet("Cables")
    ws.sheet_view.showGridLines = False

    headers = [
        "Tag", "Size", "Conductor", "Insulation",
        "Voltage (kV)", "Length (m)", "Base Ampacity (A)",
        "Derating Factor", "Derated Ampacity (A)",
    ]
    _header_style(ws, 1, len(headers), "Cable Schedule")
    _col_header_row(ws, 2, headers)

    for i, c in enumerate(cables):
        s = c.summary()
        _data_row(ws, i + 3, [
            s["id"], s["size"], s["conductor"], s["insulation"],
            s["voltage_kv"], s["length_m"], s["base_ampacity_A"],
            s["derating_factor"], s["derated_ampacity_A"],
        ], alt=bool(i % 2))

    _auto_width(ws)


def _build_breaker_sheet(wb: Workbook, breakers: list[Breaker]) -> None:
    ws = wb.create_sheet("Breakers")
    ws.sheet_view.showGridLines = False

    headers = [
        "Tag", "Type", "Voltage (kV)", "Rating (A)",
        "Interrupting (kA)", "Trip Time (s)",
    ]
    _header_style(ws, 1, len(headers), "Circuit Breaker Schedule")
    _col_header_row(ws, 2, headers)

    for i, b in enumerate(breakers):
        s = b.summary()
        _data_row(ws, i + 3, [
            s["id"], s["type"], s["voltage_kv"], s["rating_A"],
            s["interrupting_kA"], s["trip_time_s"],
        ], alt=bool(i % 2))

    _auto_width(ws)


def _build_transformer_sheet(wb: Workbook, transformers: list[Transformer]) -> None:
    ws = wb.create_sheet("Transformers")
    ws.sheet_view.showGridLines = False

    headers = [
        "Tag", "kVA", "Primary (kV)", "Secondary (kV)",
        "Z%", "Tap", "Vector Group", "Cooling",
        "Turns Ratio", "FLC Pri (A)", "FLC Sec (A)", "Z (Ω sec)",
    ]
    _header_style(ws, 1, len(headers), "Transformer Schedule")
    _col_header_row(ws, 2, headers)

    for i, t in enumerate(transformers):
        s = t.summary()
        _data_row(ws, i + 3, [
            s["id"], s["kva"], s["primary_kv"], s["secondary_kv"],
            s["impedance_pct"], s["tap_position"], s["vector_group"], s["cooling"],
            s["turns_ratio"], s["flc_primary_A"], s["flc_secondary_A"], s["Z_secondary_ohm"],
        ], alt=bool(i % 2))

    _auto_width(ws)


def _build_summary_sheet(
    wb: Workbook,
    cables: list[Cable],
    breakers: list[Breaker],
    transformers: list[Transformer],
    project_title: str,
) -> None:
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    # Title block
    ws.merge_cells("A1:B1")
    ws["A1"] = project_title
    ws["A1"].font = Font(bold=True, size=16, color=WHITE, name="Calibri")
    ws["A1"].fill = PatternFill("solid", fgColor=DARK_BLUE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:B2")
    ws["A2"] = f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, color="FF404040", name="Calibri", size=9)
    ws["A2"].fill = PatternFill("solid", fgColor=LIGHT_GREY)
    ws["A2"].alignment = Alignment(horizontal="center")

    items = [
        ("Equipment Type", "Count"),
        ("Cables",       len(cables)),
        ("Breakers",     len(breakers)),
        ("Transformers", len(transformers)),
        ("Total Items",  len(cables) + len(breakers) + len(transformers)),
    ]
    for row_offset, (label, value) in enumerate(items, start=4):
        ws.cell(row=row_offset, column=1, value=label).font = Font(
            bold=(label in ("Equipment Type", "Total Items")),
            name="Calibri", size=10,
        )
        ws.cell(row=row_offset, column=2, value=value).font = Font(name="Calibri", size=10)
        if label == "Total Items":
            for col in (1, 2):
                ws.cell(row=row_offset, column=col).fill = PatternFill("solid", fgColor=LIGHT_BLUE)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_equipment_report(
    equipment_list: list[Union[Cable, Breaker, Transformer]],
    output_path: str | Path,
    project_title: str = "Electrical Equipment Schedule",
) -> Path:
    """
    Generate a formatted Excel equipment report.

    Parameters
    ----------
    equipment_list : list
        Any mix of ``Cable``, ``Breaker``, and ``Transformer`` instances.
    output_path : str or Path
        Destination path for the .xlsx file.
    project_title : str
        Title shown on the summary sheet.

    Returns
    -------
    Path
        Resolved path to the written workbook.

    Example
    -------
    >>> from etap.equipment import Cable, Breaker, Transformer
    >>> from etap.report_generator import generate_equipment_report
    >>> items = [
    ...     Cable("CB-001", "350 kcmil", 120, 0.48, 310, 0.8),
    ...     Breaker("52-MCC1", 400, 22, 0.033),
    ...     Transformer("T-1", 1000, 13.8, 0.48, 5.75),
    ... ]
    >>> path = generate_equipment_report(items, "output/equipment_report.xlsx")
    """
    cables       = [e for e in equipment_list if isinstance(e, Cable)]
    breakers     = [e for e in equipment_list if isinstance(e, Breaker)]
    transformers = [e for e in equipment_list if isinstance(e, Transformer)]

    wb = Workbook()
    # Remove default sheet
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    _build_summary_sheet(wb, cables, breakers, transformers, project_title)
    if cables:
        _build_cable_sheet(wb, cables)
    if breakers:
        _build_breaker_sheet(wb, breakers)
    if transformers:
        _build_transformer_sheet(wb, transformers)

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out.resolve()
